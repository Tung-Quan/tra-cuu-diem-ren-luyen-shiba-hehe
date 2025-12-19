"""
CSV reading and parsing utilities with Vietnamese text support.
"""
import io
import csv
import re
import requests
from typing import List, Optional

from backend.utils.text_processing import fix_vietnamese_text, decode_http_response
from backend.utils.url_helpers import (
    is_google_sheets_url,
    build_gsheets_csv_export,
    normalize_to_gsheets_csv_export
)


# =========================
# CSV Parsing
# =========================
def read_csv_text(csv_text: str) -> List[List[str]]:
    """
    Parse CSV text into a matrix of strings.
    Applies Vietnamese text fixing.
    """
    csv_text = fix_vietnamese_text(csv_text)
    reader = csv.reader(io.StringIO(csv_text))
    return [row for row in reader]


def read_csv_bytes(csv_bytes: bytes, encoding_hint: Optional[str] = None) -> List[List[str]]:
    """
    Parse CSV from bytes with proper encoding detection.
    
    Priority:
    1. encoding_hint if provided
    2. utf-8-sig (handles BOM)
    3. utf-8
    4. fallback with replace
    """
    if encoding_hint:
        try:
            return read_csv_text(csv_bytes.decode(encoding_hint, errors="replace"))
        except Exception:
            pass
    
    try:
        return read_csv_text(csv_bytes.decode("utf-8-sig"))
    except UnicodeDecodeError:
        try:
            return read_csv_text(csv_bytes.decode("utf-8"))
        except UnicodeDecodeError:
            return read_csv_text(csv_bytes.decode(encoding_hint or "utf-8", errors="replace"))


def csv_to_text(rows: List[List[str]]) -> str:
    """Convert matrix of rows back to CSV string."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerows(rows)
    return output.getvalue()


# =========================
# Safe CSV Fetching
# =========================
def safe_fetch_csv_text(url: str, gid: Optional[str] = None) -> tuple[Optional[str], dict]:
    """
    Safely fetch CSV text from URL (especially Google Sheets).
    
    Always normalizes to docs.google.com export format.
    Never raises - returns (text, info) with error details if failed.
    
    Returns:
        (csv_text, info) where info = {
            "normalized_url": str,
            "status": int|None,
            "error": str|None
        }
    """
    info = {"normalized_url": None, "status": None, "error": None}

    # Normalize to Google Sheets CSV export
    normalized_url, used_gid = normalize_to_gsheets_csv_export(url, gid_hint=gid)
    info["normalized_url"] = normalized_url

    try:
        resp = requests.get(normalized_url, timeout=25, headers={"Accept": "text/csv"})
        info["status"] = resp.status_code
        
        if resp.status_code >= 400:
            info["error"] = f"HTTP {resp.status_code}"
            return None, info
        
        # Decode with utf-8-sig priority
        try:
            text = resp.content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = resp.text
        
        return text, info
    
    except requests.RequestException as e:
        info["error"] = f"RequestException: {e}"
        return None, info


# =========================
# Delimiter Detection
# =========================
def guess_delimiter(lines: List[str]) -> Optional[str]:
    """
    Guess the delimiter used in tabular text.
    
    Tests: tab, comma, semicolon, pipe
    Returns delimiter with best score (most columns, most consistent)
    """
    candidates = ["\t", ",", ";", "|"]
    best = None
    best_score = -1
    
    for delim in candidates:
        splits = [len(line.split(delim)) for line in lines if line.strip()]
        if not splits:
            continue
        
        # Score: average columns - penalty for inconsistency
        avg_cols = sum(splits) / len(splits)
        inconsistency = (max(splits) - min(splits)) * 0.1
        score = avg_cols - inconsistency
        
        if score > best_score and max(splits) >= 2:
            best_score = score
            best = delim
    
    if best:
        return best
    
    # Fallback: check for multiple consecutive spaces
    spaced = sum(1 for line in lines[:20] if re.search(r"\s{2,}", line))
    return r"\s{2,}" if spaced >= max(1, len(lines[:20]) // 4) else None


def parse_plaintext_as_table(text: str, max_rows: int = 5) -> List[List[str]]:
    """
    Parse plain text as a table by guessing delimiter.
    """
    lines = [ln.strip() for ln in (text or "").splitlines() if ln.strip()]
    if not lines:
        return []
    
    delim = guess_delimiter(lines[:100])
    rows: List[List[str]] = []
    
    for ln in lines[:max_rows]:
        if delim == r"\s{2,}":
            parts = re.split(r"\s{2,}", ln)
        elif delim:
            parts = ln.split(delim)
        else:
            parts = [ln]
        rows.append([p.strip() for p in parts])
    
    return rows


# =========================
# Excel Support
# =========================
def xlsx_bytes_to_csvtext(data: bytes) -> Optional[str]:
    """Convert Excel .xlsx bytes to CSV text."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        out = io.StringIO()
        w = csv.writer(out)
        
        for ws in wb.worksheets:
            w.writerow([f"-- SHEET: {ws.title} --"])
            for row in ws.iter_rows(values_only=True):
                w.writerow(["" if v is None else str(v) for v in row])
        
        return out.getvalue()
    except Exception as e:
        from backend.config import debug_log
        debug_log(f"[excel] parse xlsx failed: {e}")
        return None


def rows_from_xlsx_bytes(data: bytes, max_rows: int = 5) -> dict[str, List[List[str]]]:
    """Parse Excel .xlsx bytes into dict of sheet_name -> rows."""
    out: dict[str, List[List[str]]] = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        
        for ws in wb.worksheets:
            rows: List[List[str]] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= max_rows:
                    break
                rows.append(["" if v is None else str(v) for v in row])
            out[ws.title] = rows
    
    except Exception as e:
        from backend.config import debug_log
        debug_log(f"[excel] rows_from_xlsx_bytes failed: {e}")
    
    return out
