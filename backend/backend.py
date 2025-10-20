# backend.py — full version (deep scan + vi-fold search + Sheets/Drive fallbacks + debug)
import os
import re
import io
import csv
import json
import time
import unicodedata
from typing import Any, Dict, List, Optional, Tuple

from fastapi import FastAPI, Query
from fastapi.middleware.cors import CORSMiddleware

from ftfy import fix_text, fix_encoding
from rapidfuzz import fuzz

# HTTP fetch & HTML parse
import httpx
from bs4 import BeautifulSoup
from urllib.parse import urlparse, parse_qs

# ---- Google Sheets via gspread (values/formulas) ----
try:
    import gspread
    from google.oauth2.service_account import Credentials
    HAS_GSPREAD = True
except Exception:
    HAS_GSPREAD = False
    Credentials = None  # type: ignore

# ---- Google APIs (Sheets + Drive) for deep scan & export ----
try:
    from googleapiclient.discovery import build
    from googleapiclient.errors import HttpError
    from googleapiclient.http import MediaIoBaseDownload
    HAS_GOOGLE_API = True
except Exception:
    HAS_GOOGLE_API = False

# ---- MySQL Database Integration ----
try:
    from backend import db_mysql
    HAS_MYSQL = True
except Exception as e:
    try:
        import db_mysql
        HAS_MYSQL = True
    except Exception as e2:
        HAS_MYSQL = False
        db_mysql = None  # type: ignore
        print(f"[WARNING] MySQL module not available: {e2}")

# =========================
# FastAPI app & CORS
# =========================
app = FastAPI(title="Link-aware Search Backend", version="2.0.0 (deep+fold+mimes)")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# =========================
# Config
# =========================
DEFAULT_SPREADSHEET_ID = os.environ.get(
    "SPREADSHEET_ID",
    "1-ypUyKglUjblgy1Gy0gITcdHF4YLdJnaCNKM_6_fCrI"  # đổi ID nếu cần
)
GOOGLE_CREDS = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS", "credentials.json")
USE_DEEP = os.environ.get("DEEP_INDEX", "1").lower() in ("1", "true", "yes")

# =========================
# Globals (in-memory DB)
# =========================
DATABASE_ROWS: List[Dict[str, Any]] = []   # row records for quick search
SHEETS: List[str] = []                     # sheet titles
LINK_POOL: Dict[str, List[Dict[str, Any]]] = {}   # map: url -> list loc (giữ tương thích)
LINK_POOL_LIST: List[Dict[str, Any]] = []         # list phẳng: mỗi phần tử là 1 link+vị trí
LINK_POOL_MAP: Dict[str, List[Dict[str, Any]]] = {}
import re
from urllib.parse import urlparse, parse_qs

_GSHEETS_ID_PATTERNS = [
    r"/spreadsheets/d/([a-zA-Z0-9-_]+)",    # docs.google.com/spreadsheets/d/<id>
    r"/file/d/([a-zA-Z0-9-_]+)",            # drive.google.com/file/d/<id>
    r"[?&]id=([a-zA-Z0-9-_]+)",             # ...?id=<id>
    r"/export/.*/\*/([a-zA-Z0-9-_]+)",      # googleusercontent.com/export/.../*/<id>
]

def extract_gsheets_file_id(u: str) -> str | None:
    for pat in _GSHEETS_ID_PATTERNS:
        m = re.search(pat, u)
        if m:
            return m.group(1)
    return None

def extract_gid(u: str) -> str | None:
    try:
        p = urlparse(u)
        qgid = (parse_qs(p.query or "").get("gid") or [None])[0]
        if qgid: return qgid
        if p.fragment:
            m = re.search(r"gid=(\d+)", p.fragment)
            if m: return m.group(1)
    except Exception:
        pass
    return None
import requests

def _safe_fetch_csv_text(u: str, gid: str | None = None) -> tuple[str | None, dict]:
    """
    Cố gắng trả về CSV text từ URL bất kỳ liên quan tới Sheets/CSV.
    Luôn chuẩn hóa về docs.google.com export. Không raise; trả (text, info).
    info = {"normalized_url":..., "status":200|..., "error":"..."} nếu có.
    """
    info = {"normalized_url": None, "status": None, "error": None}

    # Nếu là sheets (hoặc các biến thể), ép về export docs.google.com
    nu, used_gid = normalize_to_gsheets_csv_export(u, gid_hint=gid)
    info["normalized_url"] = nu

    try:
        resp = requests.get(nu, timeout=25, headers={"Accept": "text/csv"})
        info["status"] = resp.status_code
        if resp.status_code >= 400:
            info["error"] = f"HTTP {resp.status_code}"
            return None, info
        # Ưu tiên utf-8-sig
        try:
            text = resp.content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = resp.text
        return text, info
    except requests.RequestException as e:
        info["error"] = f"RequestException: {e}"
        return None, info


def normalize_to_gsheets_csv_export(u: str, gid_hint: str | None = None) -> tuple[str, str | None]:
    """
    Trả về (export_url, gid_used). Luôn ép về docs.google.com/spreadsheets/d/<id>/export?format=csv&gid=<gid>
    """
    file_id = extract_gsheets_file_id(u)
    if not file_id:
        return u, gid_hint  # không nhận ra -> trả về nguyên vẹn

    gid = gid_hint or extract_gid(u) or "0"
    export_url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=csv&gid={gid}"
    return export_url, gid

# NOTE: `add_link` removed — it was unused and incorrectly mutated LINK_POOL (dict).
# Use LINK_POOL.setdefault(url, []) and LINK_POOL_LIST append sites during indexing instead.
# Debug ring buffer for logs
DEBUG_LOG: List[str] = []
DEBUG_LOG_MAX = 5000
STATS: Dict[str, Any] = {"built_at": None, "total_links": 0, "per_sheet": []}

# Skip / cache
GSPREAD_SKIP_IDS: set[str] = set()        # spreadsheet IDs not suitable for gspread
URL_ACCESS_CACHE: dict[str, str] = {}     # url -> "ok"|"private"|"unsupported"
DRIVE_META_CACHE: dict[str, dict] = {}    # fileId -> {id,name,mimeType}
# =========================
# --- VN-safe decoding helpers ----------------------------------------------
# =========================

# Dấu hiệu thường gặp của mojibake khi UTF-8 bị đọc sai
_MOJIBAKE_SIGNS = ("Ã", "Â", "Æ°", "Ä‘", "áº", "á»", "â€", "Ê", "Ð", "Þ")

def _looks_mojibake(s: str) -> bool:
    if not s:
        return False
    # có ký tự dạng replacement hoặc các cụm đặc trưng
    return any(x in s for x in _MOJIBAKE_SIGNS) or "\ufffd" in s  # �

def _latin1_to_utf8(s: str) -> Optional[str]:
    try:
        return s.encode("latin-1", errors="strict").decode("utf-8", errors="strict")
    except Exception:
        return None

def _cp1252_to_utf8(s: str) -> Optional[str]:
    try:
        return s.encode("cp1252", errors="strict").decode("utf-8", errors="strict")
    except Exception:
        return None

def vn_strict_fix(s: str) -> str:
    """
    Sửa tiếng Việt *mạnh tay* cho các chuỗi đã bị dán sai mã:
    1) ftfy.fix_encoding
    2) latin-1 -> utf-8 (nếu còn lỗi dấu)
    3) cp1252  -> utf-8 (nếu còn)
    4) ftfy.fix_text + chuẩn hoá Unicode NFC
    """
    if s is None:
        return ""
    s0 = str(s)

    # Bước 1: thử fix_encoding
    try:
        s1 = fix_encoding(s0)
    except Exception:
        s1 = s0

    # Nếu vẫn còn dấu hiệu mojibake, thử latin-1 -> utf-8
    if _looks_mojibake(s1):
        s2 = _latin1_to_utf8(s1) or s1
    else:
        s2 = s1

    # Nếu vẫn còn, thử cp1252 -> utf-8
    if _looks_mojibake(s2):
        s3 = _cp1252_to_utf8(s2) or s2
    else:
        s3 = s2

    # Bước cuối: dọn dẹp ký tự lạ + chuẩn hoá Unicode
    try:
        s4 = fix_text(s3)
    except Exception:
        s4 = s3

    s4 = s4.replace("\u00A0", " ").replace("\u200b", "")
    s4 = re.sub(r"[ \t]+", " ", s4)
    s4 = re.sub(r"\n{3,}", "\n\n", s4)
    s4 = unicodedata.normalize("NFC", s4)

    return s4.strip()

def vn_fix_if_needed(s: str) -> str:
    """
    Sửa tiếng Việt ngay từ đầu:
    - Nếu phát hiện các pattern mojibake -> dùng fix_encoding (mạnh nhất cho UTF-8 bị decode sai)
    - Sau đó chạy fix_text để dọn ký tự lạ/khoảng trắng.
    - Không 'ép' lại encode/decode lần nữa để tránh hỏng chuỗi đã đúng.
    """
    try:
        # Chỉ gọi fix_encoding khi thực sự thấy dấu hiệu mojibake để tránh overfix
        if any(sign in s for sign in _MOJIBAKE_SIGNS):
            from ftfy import fix_encoding  # import nội bộ để cảm giác nhẹ ký sinh
            s = fix_encoding(s)
        s = fix_text(s)
    except Exception:
        pass
    # Chuẩn hóa khoảng trắng phổ biến
    s = s.replace("\u00A0", " ").replace("\u200b", "")  # NBSP, zero-width space
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()

def http_decode_bytes(resp) -> str:
    """
    Giải mã bytes HTTP theo UTF-8 trước, fallback sang encoding server báo,
    rồi cuối cùng mới đến 'replace'. Không dùng .text ngay để tránh decode sai sớm.
    """
    # 1) Ưu tiên utf-8-sig (lo BOM) -> utf-8
    try:
        return resp.content.decode("utf-8-sig")
    except UnicodeDecodeError:
        pass

    # 2) Thử encoding mà server gợi ý / chẩn đoán
    enc: Optional[str] = None
    try:
        enc = getattr(resp, "encoding", None)
        if not enc:
            enc = getattr(resp, "apparent_encoding", None)
    except Exception:
        enc = None

    if enc:
        try:
            return resp.content.decode(enc, errors="replace")
        except Exception:
            pass

    # 3) Cùng lắm mới dùng requests .text (đã có fallback nội bộ)
    try:
        return resp.text
    except Exception:
        # 4) Chốt hạ: giải mã kiểu 'replace' để không vỡ flow
        return resp.content.decode("utf-8", errors="replace")


# --- CSV/Excel safe readers -------------------------------------------------
def read_csv_text(csv_text: str) -> list[list[str]]:
    """
    Đọc CSV text (đã là str) thành ma trận; giữ nguyên tiếng Việt.
    """
    csv_text = vn_fix_if_needed(csv_text)
    reader = csv.reader(io.StringIO(csv_text))
    return [row for row in reader]

def read_csv_bytes(csv_bytes: bytes, encoding_hint: Optional[str] = None) -> list[list[str]]:
    """
    Đọc CSV từ bytes; ưu tiên utf-8-sig -> utf-8 -> hint -> replace.
    """
    if encoding_hint:
        try:
            return read_csv_text(csv_bytes.decode(encoding_hint, errors="replace"))
        except Exception:
            pass
    try:
        return read_csv_text(csv_bytes.decode("utf-8-sig"))
    except UnicodeDecodeError:
        try:
            return read_csv_text(csv_bytes.decode("utf-8"))
        except UnicodeDecodeError:
            return read_csv_text(csv_bytes.decode(encoding_hint or "utf-8", errors="replace"))

def normalize_vn_text_block(text: str) -> str:
    """
    Dùng cho HTML/Docs/Sheets export: gom text về dạng sạch tiếng Việt chuẩn.
    Gọi cái này NGAY SAU khi bạn đã có text gốc.
    """
    return vn_fix_if_needed(text)


# =========================
# Logging helper
# =========================
def _dlog(msg: str):
    ts = time.strftime("%Y-%m-%d %H:%M:%S")
    line = f"[{ts}] {msg}"
    print(line, flush=True)
    DEBUG_LOG.append(line)
    if len(DEBUG_LOG) > DEBUG_LOG_MAX:
        del DEBUG_LOG[: len(DEBUG_LOG) - DEBUG_LOG_MAX]

# =========================
# Utilities
# =========================
COL_LETTERS = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
def _a1_addr(row: int, col: int) -> str:
    c = col
    label = ""
    while c:
        c, rem = divmod(c - 1, 26)
        label = COL_LETTERS[rem] + label
    return f"{label}{row}"

URL_RE = re.compile(r"https?://[^\s)>\]\"']+", re.IGNORECASE)
DOCS_ID_RE = re.compile(r"docs\.google\.com/document/d/([a-zA-Z0-9_-]+)")
SHEETS_ID_RE = re.compile(r"docs\.google\.com/spreadsheets/d/([a-zA-Z0-9_-]+)")
GID_RE = re.compile(r"[#?&]gid=(\d+)")

# NOTE: duplicate implementation of `_iter_all_indexed_links` removed here.
# The fuller implementation appears later in the file and is used by search/scan logic.



def classify_google_url(url: str) -> Tuple[str, Optional[str], Optional[str]]:
    """
    Return (kind, file_id, gid?) where kind in {'docs','sheets','unknown'}.
    """
    m = DOCS_ID_RE.search(url)
    if m:
        return ("docs", m.group(1), None)
    m = SHEETS_ID_RE.search(url)
    if m:
        gid = None
        g = GID_RE.search(url)
        if g:
            gid = g.group(1)
        return ("sheets", m.group(1), gid)
    return ("unknown", None, None)

def export_candidates(url: str) -> List[str]:
    """
    Possible export endpoints for public Sheets/Docs.
    Sheets: CSV (with gid if present)
    Docs: TXT
    """
    kind, file_id, maybe_gid = classify_google_url(url)
    if kind == "docs" and file_id:
        return [f"https://docs.google.com/document/d/{file_id}/export?format=txt"]
    if kind == "sheets" and file_id:
        base = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=csv"
        return [f"{base}&gid={maybe_gid}"] if maybe_gid else [base]
    return [url]

def extract_links_from_cell(val: Any, formula: Any) -> List[str]:
    """
    Pull URLs from cell value and formula (HYPERLINK). De-duplicate by base.
    """
    val_s = val if isinstance(val, str) else ("" if val is None else str(val))
    fm_s = formula if isinstance(formula, str) else ("" if formula is None else str(formula))
    links: List[str] = []
    if val_s:
        links += URL_RE.findall(val_s)
    if fm_s:
        m = re.search(r'HYPERLINK\s*\(\s*"([^"]+)"', fm_s, re.IGNORECASE)
        if m:
            u = m.group(1)
            if not u.startswith("http") and ("docs.google.com" in u or "drive.google.com" in u):
                u = "https://" + u
            links.append(u)
        links += URL_RE.findall(fm_s)
    out, seen = [], set()
    for u in links:
        if not isinstance(u, str):
            continue
        cleaned = u.strip().strip('")]}').strip()
        if not cleaned.startswith("http"):
            continue
        base = cleaned.split("#")[0].split("?")[0]
        if base not in seen:
            out.append(cleaned)
            seen.add(base)
    return out

def repair_text(s: str) -> str:
    return vn_strict_fix(s)

def fold_vi(s: str) -> str:
    """
    Chuẩn hoá để tìm không dấu:
    - Sửa mojibake (repair_text)
    - Đổi đ/Đ -> d/D
    - NFKD rồi bỏ dấu kết hợp
    """
    if not isinstance(s, str):
        s = "" if s is None else str(s)
    s = repair_text(s)
    s = s.replace("đ", "d").replace("Đ", "D")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s.lower()

def normalize_query(q: str) -> Tuple[str, str]:
    """Trả về (q_fixed, q_fold) = (đã sửa font, đã bỏ dấu)."""
    q_fixed = repair_text(q or "")
    q_fold = fold_vi(q_fixed)
    return q_fixed, q_fold


# =========================
# Clients & helpers (Google)
# =========================
def _resolve_creds_path(p: str) -> str:
    if not os.path.isabs(p) and not os.path.exists(p):
        cand = os.path.join(os.path.dirname(__file__), p)
        if os.path.exists(cand):
            return cand
    return p

def _gspread_client() -> Optional["gspread.Client"]:
    if not HAS_GSPREAD or Credentials is None:
        return None
    creds_path = _resolve_creds_path(GOOGLE_CREDS)
    try:
        creds = Credentials.from_service_account_file(
            creds_path,
            scopes=[
                "https://www.googleapis.com/auth/spreadsheets.readonly",
                "https://www.googleapis.com/auth/drive.readonly",
            ],
        )
        return gspread.authorize(creds)
    except Exception as e:
        _dlog(f"[gspread] cannot authorize: {e}")
        return None

_SHEETS_SERVICE = None
def _sheets_service():
    global _SHEETS_SERVICE
    if _SHEETS_SERVICE:
        return _SHEETS_SERVICE
    if not HAS_GOOGLE_API or Credentials is None:
        _dlog("[deep] googleapiclient not available")
        return None
    try:
        creds = Credentials.from_service_account_file(
            _resolve_creds_path(GOOGLE_CREDS),
            scopes=[
                "https://www.googleapis.com/auth/spreadsheets.readonly",
                "https://www.googleapis.com/auth/drive.readonly",
            ],
        )
        _SHEETS_SERVICE = build("sheets", "v4", credentials=creds, cache_discovery=False)
        return _SHEETS_SERVICE
    except Exception as e:
        _dlog(f"[deep] cannot build sheets service: {e}")
        return None

_DRIVE_SERVICE = None
def _drive_service():
    global _DRIVE_SERVICE
    if _DRIVE_SERVICE:
        return _DRIVE_SERVICE
    if not HAS_GOOGLE_API or Credentials is None:
        _dlog("[drive] googleapiclient not available")
        return None
    try:
        creds = Credentials.from_service_account_file(
            _resolve_creds_path(GOOGLE_CREDS),
            scopes=["https://www.googleapis.com/auth/drive.readonly"],
        )
        _DRIVE_SERVICE = build("drive", "v3", credentials=creds, cache_discovery=False)
        return _DRIVE_SERVICE
    except Exception as e:
        _dlog(f"[drive] cannot build service: {e}")
        return None

def _drive_get_meta(file_id: str) -> Optional[dict]:
    if not file_id:
        return None
    if file_id in DRIVE_META_CACHE:
        return DRIVE_META_CACHE[file_id]
    svc = _drive_service()
    if not svc:
        return None
    try:
        meta = svc.files().get(
            fileId=file_id,
            fields="id,name,mimeType,shortcutDetails"
        ).execute()
        # resolve shortcut -> target
        if (meta.get("mimeType") == "application/vnd.google-apps.shortcut" and
            (meta.get("shortcutDetails") or {}).get("targetId")):
            target_id = meta["shortcutDetails"]["targetId"]
            meta2 = svc.files().get(
                fileId=target_id,
                fields="id,name,mimeType"
            ).execute()
            meta = {"id": meta2.get("id"), "name": meta2.get("name"), "mimeType": meta2.get("mimeType")}
        DRIVE_META_CACHE[file_id] = meta
        return meta
    except Exception as e:
        _dlog(f"[drive] get meta failed {file_id}: {e}")
        return None

def _drive_download_bytes(file_id: str) -> Optional[bytes]:
    svc = _drive_service()
    if not svc:
        return None
    try:
        req = svc.files().get_media(fileId=file_id)
        buf = io.BytesIO()
        downloader = MediaIoBaseDownload(buf, req)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        return buf.getvalue()
    except Exception as e:
        _dlog(f"[drive] download failed {file_id}: {e}")
        return None


def _extract_gid_from_url(u: str) -> str | None:
    try:
        p = urlparse(u)
        qgid = (parse_qs(p.query or "").get("gid") or [None])[0]
        if qgid:
            return qgid
        if p.fragment:
            m = re.search(r"gid=(\d+)", p.fragment)
            if m:
                return m.group(1)
    except Exception:
        pass
    return None

def _coerce_link_item(item) -> dict | None:
    """
    Chuẩn hoá 1 phần tử bất kỳ trong kho link về dạng:
    {"url": str, "sheet": str, "row": str|int, "gid": str|None}
    Trả về None nếu không trích được URL.
    """
    # 1) URL là string
    if isinstance(item, str):
        u = item.strip()
        return {"url": u, "sheet": "", "row": "", "gid": _extract_gid_from_url(u)}

    # 2) dict
    if isinstance(item, dict):
        # a) dạng phẳng
        u = item.get("url") or item.get("href") or item.get("link")
        if u:
            return {
                "url": str(u).strip(),
                "sheet": str(item.get("sheet") or item.get("sheet_name") or ""),
                "row": item.get("row") or item.get("a1") or "",
                "gid": item.get("gid") or _extract_gid_from_url(str(u)),
            }
        # b) dict-of-dicts: {"<key>": {"url": ...}, ...}
        #    lấy hết các value con có "url"
        sub = []
        for v in item.values():
            if isinstance(v, dict):
                u2 = v.get("url") or v.get("href") or v.get("link")
                if u2:
                    sub.append({
                        "url": str(u2).strip(),
                        "sheet": str(v.get("sheet") or v.get("sheet_name") or ""),
                        "row": v.get("row") or v.get("a1") or "",
                        "gid": v.get("gid") or _extract_gid_from_url(str(u2)),
                    })
        if sub:
            # trả về dấu hiệu đặc biệt để caller unpack tiếp
            return {"__many__": sub}

    # 3) tuple/list: cố gắng lấy phần tử đầu là URL
    if isinstance(item, (list, tuple)) and item:
        u = item[0]
        if isinstance(u, str):
            # đoán cấu trúc (url, sheet, row, gid) hoặc (url, gid)
            sheet = ""
            row = ""
            gid = None
            if len(item) >= 4:
                sheet = str(item[1])
                row = item[2]
                gid = str(item[3]) if item[3] is not None else None
            elif len(item) == 2:
                gid = str(item[1]) if item[1] is not None else None
            return {
                "url": u.strip(),
                "sheet": sheet,
                "row": row,
                "gid": gid or _extract_gid_from_url(u),
            }

    return None

def _iter_all_indexed_links(limit: int | None = None) -> list[dict]:
    # Dùng LINK_POOL_LIST nếu bạn đã làm theo patch trước; nếu không có, fallback từ LINK_POOL_MAP
    pool = globals().get("LINK_POOL_LIST")
    if isinstance(pool, list):
        return pool if (limit is None or limit >= len(pool)) else pool[:limit]

    # Fallback từ LINK_POOL_MAP (dict: url -> [loc])
    out = []
    pool_map = globals().get("LINK_POOL", {})
    for u, locs in (pool_map or {}).items():
        for loc in (locs or []):
            out.append({
                "url": u,
                "sheet": loc.get("sheet", ""),
                "row": loc.get("row", ""),
                "col": loc.get("col", ""),
                "address": loc.get("address", ""),
                "gid": loc.get("sheet_gid") or loc.get("gid"),
                "sheet_name": loc.get("sheet_name", "")
            })
            if limit and len(out) >= limit:
                return out
    return out

def _xlsx_bytes_to_csvtext(data: bytes) -> Optional[str]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        out = io.StringIO()
        w = csv.writer(out)
        for ws in wb.worksheets:
            w.writerow([f"-- SHEET: {ws.title} --"])
            for row in ws.iter_rows(values_only=True):
                w.writerow([("" if v is None else str(v)) for v in row])
        return out.getvalue()
    except Exception as e:
        _dlog(f"[excel] parse xlsx failed: {e}")
        return None
def _sheets_values_rows_by_gid(file_id: str, gid: Optional[str], max_rows: int = 5) -> Optional[Dict[str, List[List[str]]]]:
    svc = _sheets_service()
    if not svc:
        return None
    try:
        meta = svc.spreadsheets().get(
            spreadsheetId=file_id,
            fields="sheets(properties(sheetId,title))"
        ).execute()
        sheets = (meta or {}).get("sheets", [])
        title = None
        if gid:
            for s in sheets:
                props = (s.get("properties") or {})
                if str(props.get("sheetId")) == str(gid):
                    title = props.get("title"); break
        if not title and sheets:
            title = (sheets[0].get("properties") or {}).get("title")
        if not title:
            return None
        vals = svc.spreadsheets().values().get(
            spreadsheetId=file_id,
            range=title,
            majorDimension="ROWS"
        ).execute().get("values", [])
        return {title: (vals[:max_rows] if isinstance(vals, list) else [])}
    except Exception as e:
        _dlog(f"[sheets] rows_by_gid error {file_id}: {e}")
        return None

def _rows_from_xlsx_bytes(data: bytes, max_rows: int = 5) -> Dict[str, List[List[str]]]:
    out: Dict[str, List[List[str]]] = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        for ws in wb.worksheets:
            rows: List[List[str]] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= max_rows:
                    break
                rows.append([("" if v is None else str(v)) for v in row])
            out[ws.title] = rows
    except Exception as e:
        _dlog(f"[excel] rows_from_xlsx_bytes failed: {e}")
    return out

def _guess_delimiter(lines: List[str]) -> Optional[str]:
    # thử theo thứ tự phổ biến
    cands = ["\t", ",", ";", "|"]
    best = None; best_score = -1
    for d in cands:
        splits = [len([*l.split(d)]) for l in lines if l.strip()]
        if not splits: 
            continue
        # ưu tiên nhiều cột và ổn định
        score = (sum(splits) / len(splits)) - (max(splits) - min(splits)) * 0.1
        if score > best_score and max(splits) >= 2:
            best_score = score; best = d
    if best:
        return best
    # fallback: nhiều khoảng trắng liên tiếp
    spaced = 0
    for l in lines[:20]:
        if re.search(r"\s{2,}", l):
            spaced += 1
    return r"\s{2,}" if spaced >= max(1, len(lines[:20])//4) else None

def _rows_from_plaintext_as_table(text: str, max_rows: int = 5) -> List[List[str]]:
    lines = [ln.strip() for ln in (text or "").splitlines() if ln.strip()]
    if not lines:
        return []
    delim = _guess_delimiter(lines[:100])  # chỉ nhìn 100 dòng đầu
    rows: List[List[str]] = []
    for ln in lines[:max_rows]:
        if delim == r"\s{2,}":
            parts = re.split(r"\s{2,}", ln)
        elif delim:
            parts = ln.split(delim)
        else:
            # không đoán được bảng → đưa cả dòng vào 1 cột
            parts = [ln]
        rows.append([p.strip() for p in parts])
    return rows

async def preview_tables_from_url(url: str, max_rows: int = 5) -> Dict[str, Any]:
    """
    Preview chỉ cho dạng bảng:
      - Google Sheets: 5 hàng đầu của sheet theo gid (ưu tiên Drive API; nếu không, dùng CSV export)
      - Excel (.xlsx / .xls) trên Drive
      - CSV (fallback)
    BỎ QUA Google Docs.
    """
    try:
        kind, file_id, maybe_gid = classify_google_url(url)
    except Exception:
        kind, file_id, maybe_gid = None, None, None

    meta = _drive_get_meta(file_id) if file_id else None
    mime = (meta or {}).get("mimeType", "")

    # -----------------------------
    # 1) GOOGLE SHEETS (qua Drive API)
    # -----------------------------
    if mime == "application/vnd.google-apps.spreadsheet":
        try:
            rows_map = _sheets_values_rows_by_gid(file_id, maybe_gid, max_rows=max_rows) or {}
            # chuẩn hoá tiếng Việt theo ô
            fixed_tables = []
            for sheet_name, rows in rows_map.items():
                fixed_rows = [
                    [vn_strict_fix(c) if c is not None else "" for c in (row or [])]
                    for row in (rows or [])[:max_rows]
                ]
                fixed_tables.append({"name": sheet_name, "rows": fixed_rows})
            return {"kind": "sheets", "tables": fixed_tables}
        except Exception as e:
            _dlog(f"[sheets] preview via Drive API failed {file_id}: {e}")
            # rơi xuống nhánh public CSV nếu có URL hợp lệ

    # ---------------------------------------
    # 2) GOOGLE SHEETS (public URL → CSV export)
    # ---------------------------------------
    try:
        if is_google_sheets_url(url):
            export_url, sheet_hint = build_gsheets_csv_export(url)
            import requests
            resp = requests.get(export_url, timeout=20)
            resp.raise_for_status()
            csv_text = http_decode_bytes(resp)       # decode utf-8/utf-8-sig chuẩn
            rows = read_csv_text(csv_text)           # CSV → ma trận
            fixed_rows = [
                [vn_strict_fix(c) if c is not None else "" for c in (row or [])]
                for row in (rows or [])[:max_rows]
            ]
            return {
                "kind": "sheets",
                "tables": [{"name": f"Google Sheets ({sheet_hint})", "rows": fixed_rows}]
            }
    except Exception as e:
        _dlog(f"[sheets] public CSV export failed: {e}")

    # ---------------------------------------
    # 3) EXCEL (.xlsx) UPLOAD TRÊN DRIVE
    # ---------------------------------------
    if mime in ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",):
        data = _drive_download_bytes(file_id) if file_id else None
        tables = []
        if data:
            try:
                rows_map = _rows_from_xlsx_bytes(data, max_rows=max_rows) or {}
                for sheet_name, rows in rows_map.items():
                    fixed_rows = [
                        [vn_strict_fix(c) if c is not None else "" for c in (row or [])]
                        for row in (rows or [])[:max_rows]
                    ]
                    tables.append({"name": sheet_name, "rows": fixed_rows})
            except Exception as e:
                _dlog(f"[excel-xlsx] preview failed: {e}")
        return {"kind": "excel", "tables": tables}

    # ---------------------------------------
    # 4) EXCEL (.xls) CŨ – nếu có
    # ---------------------------------------
    if mime in ("application/vnd.ms-excel",):
        try:
            import pandas as pd  # cần xlrd<=1.2.0 cho .xls
            data = _drive_download_bytes(file_id) if file_id else None
            if data:
                import io
                with io.BytesIO(data) as f:
                    df = pd.read_excel(f, sheet_name=0, engine=None)
                rows = []
                n = min(len(df), max_rows)
                for i in range(n):
                    row_vals = df.iloc[i].tolist()
                    rows.append([
                        "" if (isinstance(v, float) and pd.isna(v)) else normalize_vn_text_block(str(v))
                        for v in row_vals
                    ])
                return {"kind": "excel_xls", "tables": [{"name": "Sheet1", "rows": rows}]}
        except Exception as e:
            _dlog(f"[excel-xls] preview failed: {e}")
        return {"kind": "excel_xls", "tables": []}

    # ---------------------------------------
    # 5) FALLBACK: CHỈ NHẬN CSV TỪ EXPORT CANDIDATES (KHÔNG PARSE DOCS/HTML)
    # ---------------------------------------
    try:
        for cand in export_candidates(url) or []:
            # chỉ chấp nhận CSV
            if not (cand.endswith(".csv") or "format=csv" in cand):
                continue
            txt = await _http_get_text(cand)
            if not txt:
                continue
            rows = read_csv_text(txt)
            fixed_rows = [
                [normalize_vn_text_block(str(c)) if c is not None else "" for c in (row or [])]
                for row in (rows or [])[:max_rows]
            ]
            return {"kind": "csv", "tables": [{"name": "Data", "rows": fixed_rows}]}
    except Exception as e:
        _dlog(f"[fallback-csv] preview failed: {e}")

    # ---------------------------------------
    # 6) Không nhận diện/không hỗ trợ
    # ---------------------------------------
    return {"kind": "unknown", "tables": []}


# =========================
# HTTP fetch helper
# =========================
async def _http_get_text(url: str, timeout: float = 20.0) -> Optional[str]:
    """
    Tải nội dung qua HTTP:
      - text/plain, text/csv: trả thẳng text
      - HTML: bóc sạch tag, trả text
    """
    try:
        async with httpx.AsyncClient(
            follow_redirects=True,
            timeout=timeout,
            headers={"User-Agent": "Mozilla/5.0"}
        ) as client:
            r = await client.get(url)
            if r.status_code != 200:
                try:
                    _dlog(f"[http] {r.status_code} on {url}")
                except Exception:
                    pass
                return None
            ct = (r.headers.get("content-type") or "").lower()
            if "text/plain" in ct or url.lower().endswith(".txt"):
                return r.text
            if "text/csv" in ct or url.lower().endswith(".csv"):
                return r.text
            html = r.text
            soup = BeautifulSoup(html, "html.parser")
            for tag in soup(["script", "style", "noscript"]):
                tag.extract()
            return soup.get_text("\n")
    except Exception as e:
        try:
            _dlog(f"[http] error {e} on {url}")
        except Exception:
            pass
        return None

def _csv_to_text(rows: List[List[str]]) -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerows(rows)
    return output.getvalue()

# =========================
# Fetch text from URL (Sheets/Docs/Excel/HTML)
# =========================
def is_google_sheets_url(u: str) -> bool:
    try:
        p = urlparse(u)
        return ("docs.google.com" in p.netloc) and ("/spreadsheets/" in p.path)
    except Exception:
        return False
    import re
from urllib.parse import urlparse

def is_google_sheets(u: str) -> bool:
    try:
        p = urlparse(u)
        return ("docs.google.com" in p.netloc and "/spreadsheets/" in p.path) \
            or ("googleusercontent.com" in p.netloc and "/export/" in p.path)
    except Exception:
        return False

def is_google_docs(u: str) -> bool:
    try:
        p = urlparse(u)
        return ("docs.google.com" in p.netloc and "/document/" in p.path)
    except Exception:
        return False


def build_gsheets_csv_export(u: str) -> tuple[str, str]:
    """
    Từ URL kiểu:
      https://docs.google.com/spreadsheets/d/<FILE_ID>/edit?gid=1159676852#gid=1159676852
    Trả về:
      (export_csv_url, sheet_name_hint)
    """
    p = urlparse(u)
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", p.path)
    file_id = m.group(1) if m else None

    q = parse_qs(p.query or "")
    gid_q = (q.get("gid") or [""])[0]

    # gid ưu tiên: fragment (#gid=...) nếu có
    gid_frag = ""
    if p.fragment:
      mf = re.search(r"gid=(\d+)", p.fragment)
      gid_frag = mf.group(1) if mf else ""

    gid = gid_frag or gid_q or "0"

    if not file_id:
        raise ValueError("Cannot parse Google Sheets file id")

    export_url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=csv&gid={gid}"
    # Không dễ lấy sheet name ở đây nếu không gọi API; dùng hint = f"gid {gid}"
    return export_url, f"gid {gid}"

def _mark_gspread_skip(file_id: str, reason: str):
    if not file_id:
        return
    if file_id not in GSPREAD_SKIP_IDS:
        GSPREAD_SKIP_IDS.add(file_id)
        _dlog(f"[fetch] skip gspread for {file_id} ({reason})")

def _mark_url_private(url: str, reason: str = "private"):
    if URL_ACCESS_CACHE.get(url) != "private":
        URL_ACCESS_CACHE[url] = "private"
        _dlog(f"[fetch] mark private for {url} ({reason})")

async def fetch_text_from_url(url: str) -> Optional[str]:
    kind, file_id, maybe_gid = classify_google_url(url)

    # Nếu đã biết url private → bỏ qua luôn
    if URL_ACCESS_CACHE.get(url) == "private":
        return None

    # Nếu có file_id, sniff mime qua Drive (resolve shortcut)
    meta = _drive_get_meta(file_id) if file_id else None
    if meta and meta.get("id") and file_id and meta["id"] != file_id:
        file_id = meta["id"]  # resolved target
    mime = (meta or {}).get("mimeType", "")

    # ===== Route theo MIME =====
    if mime == "application/vnd.google-apps.spreadsheet":
        # gspread trước (nếu không skip)
        if HAS_GSPREAD and (file_id not in GSPREAD_SKIP_IDS):
            client = _gspread_client()
            if client:
                try:
                    ss = client.open_by_key(file_id)
                    if maybe_gid:
                        ws = ss.get_worksheet_by_id(int(maybe_gid)) or next(
                            (w for w in ss.worksheets()
                             if str(getattr(w, "_properties", {}).get("sheetId")) == str(maybe_gid)),
                            None
                        )
                    else:
                        ws = ss.sheet1
                    if ws:
                        values = ws.get_all_values()
                        return repair_text(_csv_to_text(values))
                except Exception as e:
                    if "not supported for this document" in str(e).lower():
                        _mark_gspread_skip(file_id, "unsupported-or-noaccess")
                    else:
                        _dlog(f"[fetch] gspread sheets failed once for {file_id}: {str(e)[:160]}")
        # Sheets API values
        csv_text = _sheets_values_csv_by_gid(file_id, maybe_gid)
        if csv_text:
            return repair_text(csv_text)
        _mark_url_private(url, "sheets_api_no_access")
        return None

    if mime == "application/vnd.google-apps.document":
        # Google Docs -> export TXT
        svc = _drive_service()
        if svc:
            try:
                data = svc.files().export(fileId=file_id, mimeType="text/plain").execute()  # type: ignore
                text = data.decode("utf-8", "ignore") if isinstance(data, bytes) else str(data)
                return repair_text(text)
            except Exception as e:
                _dlog(f"[drive] export txt failed {file_id}: {e}")
        _mark_url_private(url, "drive_export_no_access")
        return None

    if mime in (
        "application/vnd.google-apps.presentation",  # Slides
        "application/vnd.google-apps.drawing",       # Drawings
        "application/vnd.google-apps.folder",        # Folder
        "application/vnd.google-apps.form",          # Forms
    ):
        _mark_url_private(url, "unsupported_mime")
        return None

    if mime in ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",):
        # Excel .xlsx uploaded to Drive
        data = _drive_download_bytes(file_id)
        if data:
            text = _xlsx_bytes_to_csvtext(data)
            if text:
                return repair_text(text)
        _mark_url_private(url, "excel_download_or_parse_failed")
        return None

    # Nếu không có/không đọc MIME (public?) → thử public export candidates
    for cand in export_candidates(url):
        text = await _http_get_text(cand)
        if text:
            return repair_text(text)

    # Cuối cùng: HTML thường
    if kind == "unknown":
        text = await _http_get_text(url)
        if text:
            return repair_text(text)

    _mark_url_private(url, "final_no_access_or_unknown_mime")
    return None

# =========================
# Deep scan helpers (rich text links in Sheets)
# =========================
def _links_from_text_runs(text_format_runs: Optional[List[Dict[str, Any]]]) -> List[str]:
    urls: List[str] = []
    for run in (text_format_runs or []):
        fmt = (run or {}).get("format") or {}
        link = fmt.get("link") or {}
        uri = link.get("uri")
        if uri:
            urls.append(uri)
    return urls

def _clean_urls(urls: List[Any]) -> List[str]:
    out, seen = [], set()
    for u in urls:
        if not isinstance(u, str):
            continue
        u = u.strip().strip('")]}').strip()
        if not u.startswith("http"):
            continue
        base = u.split("#")[0].split("?")[0]
        if base in seen:
            continue
        seen.add(base)
        out.append(u)
    return out

def deep_scan_all_sheets(service, spreadsheet_id: str) -> Dict[str, List[Dict[str, Any]]]:
    """
    Read ALL sheets with includeGridData=True and extract links from:
    - cell.hyperlink
    - userEnteredValue.formulaValue (HYPERLINK & naked http)
    - userEnteredValue.stringValue (typed http)
    - textFormatRuns.format.link.uri (rich text link)
    - formattedValue (rendered)
    - note (links in cell note)
    Return: {sheet_title: [{url,row,col,address}]}
    """
    try:
        resp = service.spreadsheets().get(
            spreadsheetId=spreadsheet_id,
            includeGridData=True,
            fields=(
                "sheets("
                "properties(title),"
                "data(rowData(values(hyperlink,userEnteredValue,formattedValue,textFormatRuns,note)))"
                ")"
            ),
        ).execute()
    except HttpError as he:
        _dlog(f"[deep] HttpError on spreadsheet: {he}")
        return {}
    except Exception as e:
        _dlog(f"[deep] error on spreadsheet: {e}")
        return {}

    results: Dict[str, List[Dict[str, Any]]] = {}
    for s in resp.get("sheets", []):
        title = (s.get("properties") or {}).get("title", "")
        data_blocks = s.get("data", []) or []
        found: List[Dict[str, Any]] = []
        for block in data_blocks:
            rowData = block.get("rowData", []) or []
            for r_idx, row in enumerate(rowData):
                values = (row or {}).get("values", []) or []
                for c_idx, cell in enumerate(values):
                    urls: List[str] = []
                    hl = cell.get("hyperlink")
                    if hl:
                        urls.append(hl)
                    uev = cell.get("userEnteredValue") or {}
                    fv = uev.get("formulaValue")
                    if fv:
                        urls.extend(URL_RE.findall(fv))
                        m = re.search(r'HYPERLINK\s*\(\s*"([^"]+)"', fv, re.IGNORECASE)
                        if m:
                            urls.append(m.group(1))
                    sv = uev.get("stringValue")
                    if sv:
                        urls.extend(URL_RE.findall(sv))
                    urls.extend(_links_from_text_runs(cell.get("textFormatRuns")))
                    fmt = cell.get("formattedValue")
                    if fmt:
                        urls.extend(URL_RE.findall(fmt))
                    note = cell.get("note")
                    if note:
                        urls.extend(URL_RE.findall(note))
                    urls = _clean_urls(urls)
                    if not urls:
                        continue
                    r1 = r_idx + 1
                    c1 = c_idx + 1
                    a1 = _a1_addr(r1, c1)
                    for u in urls:
                        found.append({"url": u, "row": r1, "col": c1, "address": a1})
        results[title] = found
    return results

# =========================
# Indexer
# =========================
def _peek_url_rows(u: str, nrows: int = 10, gid: str | None = None) -> list[list[str]]:
    """
    Đọc nhanh 10 dòng đầu từ một URL bảng (Google Sheets/CSV) và sửa tiếng Việt từng ô.
    Không raise lỗi — nếu có vấn đề sẽ trả [].
    """
    try:
        # Google Sheets -> CSV theo gid
        if is_google_sheets_url(u):
            # nếu caller truyền gid riêng, gắn vào URL gốc trước khi build export
            base = u.split("#")[0]
            if gid is not None and "gid=" not in base:
                sep = "&" if "?" in base else "?"
                u = f"{base}{sep}gid={gid}"
            export_url, _hint = build_gsheets_csv_export(u)
            import requests
            resp = requests.get(export_url, timeout=25)
            resp.raise_for_status()
            csv_text = http_decode_bytes(resp)
            rows = read_csv_text(csv_text)
            out = []
            for row in rows[:nrows]:
                out.append([vn_strict_fix("" if c is None else str(c)) for c in (row or [])])
            return out

        # CSV public
        if u.endswith(".csv") or "format=csv" in u:
            import requests
            resp = requests.get(u, timeout=25)
            resp.raise_for_status()
            csv_text = http_decode_bytes(resp)
            rows = read_csv_text(csv_text)
            out = []
            for row in rows[:nrows]:
                out.append([vn_strict_fix("" if c is None else str(c)) for c in (row or [])])
            return out
    except Exception as e:
        _dlog(f"[peek] failed {u}: {e}")
    return []

def index_sources(verbose: bool = False, deep: bool = False) -> Tuple[List[Dict[str, Any]], List[str]]:
    rows: List[Dict[str, Any]] = []
    sheets: List[str] = []

    # reset kho link
    LINK_POOL.clear()        # dict: url -> [loc]
    LINK_POOL_LIST.clear()   # list phẳng: [{url, sheet, row, col, address, gid, sheet_name}]

    STATS["per_sheet"] = []
    start_ts = time.time()

    if not DEFAULT_SPREADSHEET_ID:
        _dlog("[index] No SPREADSHEET_ID set → empty index.")
        return rows, sheets
    if not HAS_GSPREAD:
        _dlog("[index] gspread not available → cannot open spreadsheet; empty index.")
        return rows, sheets

    client = _gspread_client()
    if not client:
        _dlog("[index] cannot get gspread client")
        return rows, sheets

    try:
        ss = client.open_by_key(DEFAULT_SPREADSHEET_ID)
        worksheets = ss.worksheets()
        _dlog(f"[index] Opened spreadsheet with {len(worksheets)} sheets")
    except Exception as e:
        _dlog(f"[index] open spreadsheet failed: {e}")
        return rows, sheets

    # gid -> sheet name
    gid_to_name: Dict[str, str] = {}
    for ws in worksheets:
        props = getattr(ws, "_properties", {}) or {}
        gid = props.get("sheetId") or props.get("id")
        if gid is not None:
            gid_to_name[str(gid)] = ws.title

    # deep rich-text links
    deep_map: Dict[str, List[Dict[str, Any]]] = {}
    if deep:
        svc = _sheets_service()
        if svc:
            _dlog("[deep] ENABLED: scanning rich-text links via Sheets API…")
            deep_map = deep_scan_all_sheets(svc, DEFAULT_SPREADSHEET_ID) or {}
            total_occ = sum(len(v) for v in deep_map.values())
            _dlog(f"[deep] SUMMARY: got {total_occ} link occurrences from API")
        else:
            _dlog("[deep] DISABLED (service not available)")

    # để tránh trùng khi đổ vào LIST phẳng
    seen_flat_keys: set[str] = set()  # key = f"{url}@@{address}"

    sheet_row_base: Dict[str, int] = {}

    for ws in worksheets:
        title = ws.title
        sheets.append(title)

        try:
            values = ws.get_all_values()
        except Exception as e:
            _dlog(f"[index] cannot read values: {title}: {e}")
            continue

        nrows = len(values)
        ncols = max((len(r) for r in values), default=0)
        try:
            if nrows > 0 and ncols > 0:
                rng = f"A1:{_a1_addr(nrows, ncols)}"
                raw_formulas = ws.get(rng, value_render_option="FORMULA")
                formulas = [[cell if isinstance(cell, str) else ("" if cell is None else str(cell)) for cell in row]
                            for row in raw_formulas]
            else:
                formulas = []
        except Exception as e:
            _dlog(f"[index] cannot read formulas: {title}: {e}")
            formulas = []

        # stats
        non_empty_cells = 0
        value_http_cells = 0
        formula_http_cells = 0
        cell_with_links = 0
        list_before = len(LINK_POOL_LIST)
        map_before = len(LINK_POOL)

        if verbose:
            _dlog(f"[index] Sheet '{title}' size ≈ {nrows}x{ncols}")

        # pack rows cho search theo hàng
        base_row_index = len(rows)
        sheet_row_base[title] = base_row_index
        for r in range(nrows):
            row_vals = values[r]
            if any((cell or "").strip() for cell in row_vals):
                rows.append({
                    "sheet": title,
                    "row": r + 1,
                    "cols": row_vals,
                    "text": " ".join(x for x in row_vals if isinstance(x, str)),
                    "links": []
                })

        # duyệt cell để trích link
        for r in range(nrows):
            for c in range(ncols):
                v = values[r][c] if c < len(values[r]) else ""
                if isinstance(v, str) and v.strip():
                    non_empty_cells += 1
                fm = formulas[r][c] if (r < len(formulas) and c < len(formulas[r])) else ""
                if ("http" in (v or "")) and isinstance(v, str):
                    value_http_cells += 1
                if ("http" in (fm or "")) or ("HYPERLINK" in (fm or "")):
                    formula_http_cells += 1

                link_list = extract_links_from_cell(v, fm)
                if not link_list:
                    continue
                cell_with_links += 1

                a1 = _a1_addr(r + 1, c + 1)

                for u in link_list:
                    # enrich gid & sheet_name
                    _, __, maybe_gid = classify_google_url(u)
                    sheet_name = gid_to_name.get(str(maybe_gid), "") if maybe_gid else ""

                    # 1) map (url -> loc list) giữ tương thích cũ
                    loc = {"sheet": title, "row": r + 1, "col": c + 1, "address": a1}
                    if maybe_gid:  # giữ tên field cũ nếu bạn đang dùng "sheet_gid"
                        loc["sheet_gid"] = maybe_gid
                        if sheet_name:
                            loc["sheet_name"] = sheet_name
                    LINK_POOL.setdefault(u, []).append(loc)

                    # 2) list phẳng (phục vụ scan_all_links)
                    k = f"{u}@@{a1}"
                    if k not in seen_flat_keys:
                        seen_flat_keys.add(k)
                        LINK_POOL_LIST.append({
                            "url": u,
                            "sheet": title,
                            "row": r + 1,
                            "col": c + 1,
                            "address": a1,
                            "gid": maybe_gid,
                            "sheet_name": sheet_name
                        })

                # đẩy link vào record hàng (để follow_links hoạt động như cũ)
                try:
                    rows[base_row_index + r]["links"].extend(link_list)
                except Exception:
                    pass

        # deep rich-text links
        deep_added_list = 0
        deep_added_map = 0
        deep_occurs = 0
        if deep and title in deep_map:
            occs = deep_map.get(title) or []
            deep_occurs = len(occs)
            for item in occs:
                u = item["url"]; r1 = item["row"]; c1 = item["col"]; a1_deep = item["address"]
                _, __, maybe_gid2 = classify_google_url(u)
                sheet_name2 = gid_to_name.get(str(maybe_gid2), "") if maybe_gid2 else ""

                # map
                exists = any(loc.get("sheet")==title and loc.get("address")==a1_deep for loc in LINK_POOL.get(u, []))
                if not exists:
                    LINK_POOL.setdefault(u, []).append({
                        "sheet": title, "row": r1, "col": c1, "address": a1_deep,
                        **({"sheet_gid": maybe_gid2} if maybe_gid2 else {}),
                        **({"sheet_name": sheet_name2} if sheet_name2 else {})
                    })
                    deep_added_map += 1

                # list phẳng
                k2 = f"{u}@@{a1_deep}"
                if k2 not in seen_flat_keys:
                    seen_flat_keys.add(k2)
                    LINK_POOL_LIST.append({
                        "url": u, "sheet": title, "row": r1, "col": c1, "address": a1_deep,
                        "gid": maybe_gid2, "sheet_name": sheet_name2
                    })
                    deep_added_list += 1

                # gắn vào row record
                try:
                    rows[sheet_row_base[title] + (r1 - 1)]["links"].append(u)
                except Exception:
                    pass

        sheet_stat = {
            "sheet": title,
            "size": f"{nrows}x{ncols}",
            "non_empty_cells": non_empty_cells,
            "value_http_cells": value_http_cells,
            "formula_http_cells": formula_http_cells,
            "cells_with_links_extracted": cell_with_links,
            "added_to_list_flat": len(LINK_POOL_LIST) - list_before,
            "added_url_keys_map": len(LINK_POOL) - map_before,
            "deep_occurrences_seen": deep_occurs,
            "deep_added_to_list_flat": deep_added_list,
            "deep_added_url_keys_map": deep_added_map,
            "flat_links_total_now": len(LINK_POOL_LIST),
            "url_keys_total_now": len(LINK_POOL),
        }
        STATS["per_sheet"].append(sheet_stat)

        _dlog(
            f"[index] '{title}': non-empty={non_empty_cells}, "
            f"value_has_http={value_http_cells}, formula_has_http/HYPERLINK={formula_http_cells}, "
            f"cells_extracted_links={cell_with_links}, "
            f"add_flat={sheet_stat['added_to_list_flat']}, add_map_keys={sheet_stat['added_url_keys_map']}, "
            f"deep_occurs={deep_occurs}, deep_add_flat={deep_added_list}, deep_add_map={deep_added_map}, "
            f"total_flat={len(LINK_POOL_LIST)}, total_map_keys={len(LINK_POOL)}"
        )

    STATS["built_at"] = time.strftime("%Y-%m-%d %H:%M:%S")
    STATS["total_links_flat"] = len(LINK_POOL_LIST)
    STATS["total_links_map_keys"] = len(LINK_POOL)
    _dlog(
        f"[index] Indexed {len(LINK_POOL_LIST)} flat link records and "
        f"{len(LINK_POOL)} unique URL keys from {len(sheets)} sheets "
        f"in {round(time.time()-start_ts,2)}s"
    )
    return rows, sheets



def build_database(verbose: bool = False, deep: bool = False) -> Tuple[List[Dict[str, Any]], List[str]]:
    return index_sources(verbose=verbose, deep=deep)

# =========================
# Search helpers
# =========================
def _tokens(q_fold: str) -> list[str]:
    # tách theo khoảng trắng, bỏ token 1 ký tự hay stopword rất ngắn nếu muốn
    return [t for t in re.split(r"\s+", q_fold) if len(t) >= 2]

def _all_tokens_in_text(q_fold: str, text_fold: str) -> bool:
    # match biên từ trên text đã bỏ dấu để tránh match một phần "nguyen" trong "truongnguyen"
    toks = _tokens(q_fold)
    return all(re.search(rf"\b{re.escape(tok)}\b", text_fold) for tok in toks)

def search_rows(query: str, top_k: int = 20, fuzz_threshold: int = 85, exact: bool = False) -> List[Dict[str, Any]]:
    results: List[Tuple[int, Dict[str, Any]]] = []
    if not query or not str(query).strip():
        return []
    q_fixed, q_fold = normalize_query(query)
    tokens = _tokens(q_fold)
    
    for row in DATABASE_ROWS:
        raw = row.get("text") or ""
        text_fixed = repair_text(raw)
        text_fold  = fold_vi(text_fixed)
        if exact:
            ok = _all_tokens_in_text(q_fold, text_fold)
            score = 100 if ok else 0
        else:
            if q_fixed.lower() in text_fixed.lower() or q_fold in text_fold:
                score = 100
            else:
                score = max(
                    fuzz.partial_ratio(q_fixed, text_fixed),
                    fuzz.partial_ratio(q_fold,  text_fold)
                )
        
        if score >= (100 if exact else fuzz_threshold):
            results.append((score, {
                "sheet": row["sheet"],
                "row": row["row"],
                "snippet": _snippet(text_fixed, q_fixed),
                "snippet_nodau": _snippet(text_fold, q_fold),
                "links": sorted(set(row.get("links", []))),
                "score": score
            }))
    results.sort(key=lambda x: x[0], reverse=True)
    return [r for _, r in results[:top_k]]


def _tokens_fold(q: str) -> list[str]:
    return [t for t in re.split(r"\s+", fold_vi(q)) if t]

def _all_tokens_word_boundary(text_fold: str, toks: list[str]) -> bool:
    return all(re.search(rf"\b{re.escape(tok)}\b", text_fold) for tok in toks)

def _match_score(q_raw: str, s_raw: str, exact: bool, fuzz_threshold: int) -> int:
    q_fix, s_fix = vn_strict_fix(q_raw), vn_strict_fix(s_raw)
    q_fold, s_fold = fold_vi(q_fix), fold_vi(s_fix)
    if exact:
        return 100 if _all_tokens_word_boundary(s_fold, _tokens_fold(q_fix)) else 0
    return max(
        fuzz.partial_ratio(q_fix, s_fix),
        fuzz.partial_ratio(q_fold, s_fold)
    )

def _snippet(s: str, q: str, window: int = 60) -> str:
    s_fix, q_fix = vn_strict_fix(s), vn_strict_fix(q)
    s_fold, q_fold = fold_vi(s_fix), fold_vi(q_fix)
    m = re.search(re.escape(q_fold), s_fold, flags=re.IGNORECASE)
    if not m:
        # fallback: lấy đầu dòng
        return s_fix[:window*2]
    i = m.start()
    left = max(0, i - window)
    right = min(len(s_fix), i + window)
    return s_fix[left:right]
# ---- CORE: search trong 1 URL (Sheets/CSV) trả về list hit ----
def _search_in_one_url_core(u: str, q: str, *, gid: str | None = None,
                            exact: bool = False, fuzz_threshold: int = 85,
                            max_rows: int = 10000) -> list[dict] | dict:
    """
    Trả list hit khi OK, hoặc dict {"error": "...", "normalized_url": "..."} nếu không đọc được.
    """
    # 1) CSV text an toàn
    csv_text, finfo = _safe_fetch_csv_text(u, gid=gid)
    if not csv_text:
        return {"error": finfo.get("error") or "fetch_failed", "normalized_url": finfo.get("normalized_url")}

    # 2) Parse CSV -> rows
    try:
        rows = read_csv_text(csv_text)
    except Exception as e:
        return {"error": f"csv_parse_failed: {e}", "normalized_url": finfo.get("normalized_url")}

    # 3) Match
    hits: list[dict] = []
    for idx, row in enumerate(rows[:max_rows], start=1):
        line = " | ".join("" if c is None else str(c) for c in row)
        score = _match_score(q, line, exact, fuzz_threshold)
        if score >= (100 if exact else fuzz_threshold):
            hits.append({
                "row": idx,
                "score": score,
                "snippet": _snippet(line, q),
                "values": [vn_strict_fix("" if c is None else str(c)) for c in row],
            })
    return hits

# =========================
# API endpoints
# =========================

@app.get("/search_url")
async def search_url(
    u: str = Query(..., description="URL sheet/csv cần tìm"),
    q: str = Query(..., description="Truy vấn (VD: họ tên)"),
    gid: Optional[str] = Query(None, description="GID sheet nếu là Google Sheets"),
    exact: bool = Query(False, description="Bắt đủ token theo biên từ"),
    fuzz_threshold: int = Query(85, ge=0, le=100),
    max_rows: int = Query(1000, ge=1, le=5000),
) -> Dict[str, Any]:
    """
    Tìm trực tiếp trong 1 URL (Google Sheets/CSV) mà không phụ thuộc index.
    Trả về các dòng có match + snippet.
    """
    hits = []

    # ----- Google Sheets -> CSV (UTF-8) -----
    if is_google_sheets_url(u):
        export_url, hint = build_gsheets_csv_export(u if gid is None else f"{u.split('#')[0].split('?')[0]}?gid={gid}")
        import requests
        resp = requests.get(export_url, timeout=25)
        resp.raise_for_status()
        csv_text = http_decode_bytes(resp)
        rows = read_csv_text(csv_text)

        for idx, row in enumerate(rows[:max_rows], start=1):
            line = " | ".join("" if c is None else str(c) for c in row)
            score = _match_score(q, line, exact, fuzz_threshold)
            if score >= (100 if exact else fuzz_threshold):
                hits.append({
                    "row": idx,
                    "score": score,
                    "snippet": _snippet(line, q),
                    "values": [vn_strict_fix("" if c is None else str(c)) for c in row],
                })

        return {"url": u, "kind": "sheets", "gid": gid, "query": q, "hits": hits}

    # ----- CSV công khai (fallback) -----
    if u.endswith(".csv") or "format=csv" in u:
        import requests
        resp = requests.get(u, timeout=25)
        resp.raise_for_status()
        csv_text = http_decode_bytes(resp)
        rows = read_csv_text(csv_text)
        for idx, row in enumerate(rows[:max_rows], start=1):
            line = " | ".join("" if c is None else str(c) for c in row)
            score = _match_score(q, line, exact, fuzz_threshold)
            if score >= (100 if exact else fuzz_threshold):
                hits.append({
                    "row": idx,
                    "score": score,
                    "snippet": _snippet(line, q),
                    "values": [vn_strict_fix("" if c is None else str(c)) for c in row],
                })
        return {"url": u, "kind": "csv", "query": q, "hits": hits}

    # Chưa hỗ trợ loại khác ở đây
    return {"url": u, "kind": "unsupported", "query": q, "hits": []}
@app.get("/health")
def health():
    return {"ok": True, "rows": len(DATABASE_ROWS), "sheets": len(SHEETS), "links": len(LINK_POOL)}

@app.get("/sheets")
def list_sheets():
    return {"sheets": SHEETS}

@app.get("/debug/link_count")
def debug_link_count():
    return {"count": len(LINK_POOL)}

@app.get("/debug/links")
def debug_links(limit: int = 50):
    return {"count": len(LINK_POOL), "items": LINK_POOL[:limit]}


@app.get("/links_index")
def links_index(limit: int = 200):
    examples = []
    for i, (u, locs) in enumerate(LINK_POOL.items()):
        if i >= limit:
            break
        examples.append({"url": u, "where": locs[:5]})
    return {"total": len(LINK_POOL), "examples": examples}

@app.get("/rebuild_db")
def rebuild_db(deep: int | None = None):
    global DATABASE_ROWS, SHEETS
    resolved_deep = USE_DEEP if deep is None else bool(deep)
    t0 = time.time()
    rows, sheets = build_database(verbose=False, deep=resolved_deep)
    DATABASE_ROWS = rows
    SHEETS = sheets
    dt = round(time.time() - t0, 2)
    return {
        "status": "ok",
        "rows": len(DATABASE_ROWS),
        "sheets": len(SHEETS),
        "links": len(LINK_POOL),
        "took_s": dt,
        "stats": STATS,
        "deep_used": resolved_deep,
    }

# ---- DEBUG ----
@app.get("/debug/reindex")
def debug_reindex(verbose: int = 1, deep: int | None = None):
    global DATABASE_ROWS, SHEETS
    resolved_deep = USE_DEEP if deep is None else bool(deep)
    t0 = time.time()
    rows, sheets = build_database(verbose=bool(verbose), deep=resolved_deep)
    DATABASE_ROWS = rows
    SHEETS = sheets
    dt = round(time.time() - t0, 2)
    return {
        "status": "ok",
        "rows": len(DATABASE_ROWS),
        "sheets": len(SHEETS),
        "links": len(LINK_POOL),
        "took_s": dt,
        "stats": STATS,
        "deep_used": resolved_deep,
        "log_tail_hint": "GET /debug/scan_log?limit=300",
    }

@app.get("/debug/stats")
def debug_stats():
    return STATS

@app.get("/debug/scan_log")
def debug_scan_log(limit: int = 300):
    limit = max(1, min(limit, 2000))
    return {"lines": DEBUG_LOG[-limit:]}

@app.get("/debug/capabilities")
def debug_capabilities():
    return {
        "HAS_GSPREAD": HAS_GSPREAD,
        "HAS_GOOGLE_API": HAS_GOOGLE_API,
        "SPREADSHEET_ID": DEFAULT_SPREADSHEET_ID,
        "CREDS_PATH": GOOGLE_CREDS,
        "USE_DEEP": USE_DEEP,
    }

@app.get("/debug/gspread_skip")
def debug_gspread_skip():
    return {"count": len(GSPREAD_SKIP_IDS), "ids": list(GSPREAD_SKIP_IDS)[:50]}

@app.post("/debug/gspread_skip/clear")
def debug_gspread_skip_clear():
    GSPREAD_SKIP_IDS.clear()
    return {"status": "ok", "cleared": True}

@app.get("/debug/service_account")
def debug_service_account():
    try:
        with open(_resolve_creds_path(GOOGLE_CREDS), "r", encoding="utf-8") as f:
            return {"client_email": json.load(f).get("client_email")}
    except Exception as e:
        return {"error": str(e)}

@app.get("/debug/check_url")
async def debug_check_url(u: str):
    kind, file_id, gid = classify_google_url(u)
    can = export_candidates(u)
    status_before = URL_ACCESS_CACHE.get(u)
    text = await fetch_text_from_url(u)
    status_after = URL_ACCESS_CACHE.get(u)
    return {
        "kind": kind, "file_id": file_id, "gid": gid,
        "export_candidates": can,
        "was_marked": status_before, "now_marked": status_after,
        "fetched": bool(text),
        "snippet": (text[:180] + "…") if text else None
    }

# ---- SEARCH ----
@app.get("/search")
async def search(
    q: str = Query(..., description="Search query"),
    top_k: int = 20,
    exact: bool = False,
    fuzz_threshold: int = 85,
    # theo hành vi cũ: đi theo link ở các hàng khớp
    follow_links: bool = False,
    link_limit: int = 10,
    # quét toàn bộ link đã index
    scan_all_links: bool = False,
    link_limit_all: int = 200,
    # DEBUG mới
    scan_all_links_debug: bool = False,
    peek_rows: int = 10,
):
    q_fixed, q_fold = normalize_query(q)

    # 1) Kết quả theo hàng (giữ nguyên)
    base_hits = search_rows(
        q_fixed, top_k=top_k, fuzz_threshold=fuzz_threshold, exact=exact
    ) or []

    link_hits: list[dict] = []
    debug_scan: list[dict] = []
    seen_urls: set[str] = set()

    # --- helper nội bộ để thêm debug cho một URL ---
    def _append_debug(u: str, src_item: dict, gid: str | None):
        if not scan_all_links_debug:
            return
        try:
            norm_u, _ = normalize_to_gsheets_csv_export(u, gid_hint=gid)
        except Exception:
            norm_u = u
        peek = _peek_url_rows(u, nrows=max(1, min(int(peek_rows), 20)), gid=gid)
        debug_scan.append({
            "url": u,
            "normalized_url": norm_u,
            "source": {
                "sheet": src_item.get("sheet", ""),
                "row": src_item.get("row", ""),
                "address": src_item.get("address", ""),
                "gid": gid,
                "sheet_name": src_item.get("sheet_name", "")
            },
            "peek": peek
        })

    # 2) FOLLOW links từ hàng khớp (giữ nguyên ngữ nghĩa, thêm chống lỗi)
    if follow_links and base_hits:
        cand_urls: list[tuple[str, str | None, dict]] = []
        for h in base_hits:
            for l in (h.get("links") or [])[:link_limit]:
                if isinstance(l, str):
                    u = l; gid = None
                else:
                    u = l.get("url") or l.get("href")
                    gid = l.get("gid") or l.get("sheet_gid")
                if not u or u in seen_urls:
                    continue
                if is_google_docs(u):
                    # không follow Google Docs (chưa hỗ trợ)
                    continue
                seen_urls.add(u)
                cand_urls.append((u, gid, {"sheet": h.get("sheet",""), "row": h.get("row",""), "address": ""}))

        for u, gid, src in cand_urls:
            try:
                url_result = _search_in_one_url_core(
                    u, q, gid=gid, exact=exact, fuzz_threshold=fuzz_threshold, max_rows=10000
                )
                # lỗi fetch/parse -> chỉ debug, không crash
                if isinstance(url_result, dict) and url_result.get("error"):
                    if scan_all_links_debug:
                        norm_u, _ = normalize_to_gsheets_csv_export(u, gid_hint=gid)
                        debug_scan.append({
                            "url": u,
                            "normalized_url": norm_u,
                            "source": src,
                            "error": url_result.get("error"),
                            "peek": _peek_url_rows(u, nrows=max(1, min(int(peek_rows), 20)), gid=gid)
                        })
                    continue

                # có hit
                if url_result:
                    link_hits.append({
                        "url": u,
                        "count": len(url_result),
                        "matches": url_result[:5],
                        "source": "row_links"
                    })
                # debug (kể cả không match)
                _append_debug(u, src, gid)
            except Exception as e:
                if scan_all_links_debug:
                    debug_scan.append({
                        "url": u,
                        "source": src,
                        "error": f"exception: {e}",
                        "peek": []
                    })

    # 3) SCAN ALL LINKS (đọc từ kho link phẳng) + DEBUG
    if scan_all_links:
        all_links = _iter_all_indexed_links(limit=link_limit_all) or []
        for item in all_links:
            u = item.get("url") if isinstance(item, dict) else None
            if not u or u in seen_urls:
                continue
            if is_google_docs(u):
                # không scan Google Docs (chưa hỗ trợ)
                continue
            seen_urls.add(u)
            gid = item.get("gid") if isinstance(item, dict) else None

            try:
                url_result = _search_in_one_url_core(
                    u, q, gid=gid, exact=exact, fuzz_threshold=fuzz_threshold, max_rows=10000
                )
                # lỗi fetch/parse
                if isinstance(url_result, dict) and url_result.get("error"):
                    if scan_all_links_debug:
                        norm_u, _ = normalize_to_gsheets_csv_export(u, gid_hint=gid)
                        debug_scan.append({
                            "url": u,
                            "normalized_url": norm_u,
                            "source": {
                                "sheet": item.get("sheet",""),
                                "row": item.get("row",""),
                                "address": item.get("address",""),
                                "gid": gid,
                                "sheet_name": item.get("sheet_name","")
                            },
                            "error": url_result.get("error"),
                            "peek": _peek_url_rows(u, nrows=max(1, min(int(peek_rows), 20)), gid=gid)
                        })
                    continue

                # có hit
                if url_result:
                    link_hits.append({
                        "url": u,
                        "count": len(url_result),
                        "matches": url_result[:5],
                        "source": {
                            "sheet": item.get("sheet",""),
                            "row": item.get("row",""),
                            "address": item.get("address","")
                        }
                    })
                # debug (kể cả không match)
                _append_debug(u, item if isinstance(item, dict) else {}, gid)

            except Exception as e:
                if scan_all_links_debug:
                    debug_scan.append({
                        "url": u,
                        "source": {
                            "sheet": item.get("sheet",""),
                            "row": item.get("row",""),
                            "address": item.get("address",""),
                            "gid": gid,
                            "sheet_name": item.get("sheet_name","")
                        },
                        "error": f"exception: {e}",
                        "peek": []
                    })

    return {
        "query": q,
        "hits": base_hits,
        "link_hits": link_hits,
        "stats": {
            "rows_considered": len(base_hits),
            "urls_considered": len(seen_urls),
            "mode": {
                "follow_links": bool(follow_links),
                "scan_all_links": bool(scan_all_links),
                "debug": bool(scan_all_links_debug)
            }
        },
        **({"debug_scan": debug_scan} if scan_all_links_debug else {})
    }



# ---- PREVIEW ----
@app.get("/preview_link")
async def preview_link(u: str, max_rows: int = 20):
    """
    Trả về preview bảng:
    - Excel (xlsx/xls trên Drive) → 5 hàng đầu
    - Google Sheets → 5 hàng đầu của sheet
    - Google Docs → parse text → bảng
    """
    result = await preview_tables_from_url(u, max_rows=max_rows)
    return {"url": u, **result}

@app.get("/search_links")
def search_links(q: str = Query(..., description="substring to match in URL"), limit: int = 50):
    qlow = q.lower()
    matches = []
    for url, locs in LINK_POOL.items():
        if qlow in url.lower():
            matches.append({"url": url, "where": locs[:5], "count": len(locs)})
            if len(matches) >= limit:
                break
    return {"query": q, "matches": matches, "total_urls": len(LINK_POOL)}

# =========================
# MySQL Integration Endpoints
# =========================

@app.get("/mysql/test")
def mysql_test():
    """Kiểm tra kết nối MySQL."""
    if not HAS_MYSQL:
        return {"ok": False, "error": "MySQL module not available. Install: pip install mysql-connector-python"}
    try:
        return db_mysql.test_connection()
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.post("/mysql/init_db")
def mysql_init_db():
    """Tạo databases và tables từ schema.sql."""
    if not HAS_MYSQL:
        return {"ok": False, "error": "MySQL module not available"}
    try:
        return db_mysql.init_databases()
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.post("/mysql/sync_links")
def mysql_sync_links(clear_first: bool = False):
    """
    Đồng bộ LINK_POOL lên MySQL database.
    clear_first=True: xóa toàn bộ dữ liệu cũ trước khi sync.
    """
    if not HAS_MYSQL:
        return {"ok": False, "error": "MySQL module not available"}
    
    try:
        if clear_first:
            deleted = db_mysql.clear_links_table()
            _dlog(f"[mysql] Cleared {deleted} old links")
        
        # Chuyển LINK_POOL_LIST sang format phù hợp
        links_to_insert = []
        for item in LINK_POOL_LIST:
            links_to_insert.append({
                "url": item.get("url", ""),
                "sheet": item.get("sheet", ""),
                "row": item.get("row"),
                "col": item.get("col"),
                "address": item.get("address", ""),
                "gid": item.get("gid"),
                "sheet_name": item.get("sheet_name", ""),
            })
        
        inserted, updated = db_mysql.insert_links_batch(links_to_insert)
        new_count = db_mysql.get_links_count()
        
        return {
            "ok": True,
            "inserted": inserted,
            "updated": updated,
            "total_in_db": new_count,
            "synced_from_memory": len(LINK_POOL_LIST),
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/mysql/links/count")
def mysql_links_count():
    """Đếm số links trong MySQL."""
    if not HAS_MYSQL:
        return {"ok": False, "error": "MySQL module not available"}
    try:
        count = db_mysql.get_links_count()
        return {"ok": True, "count": count}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/mysql/links/summary")
def mysql_links_summary():
    """Thống kê links theo sheet."""
    if not HAS_MYSQL:
        return {"ok": False, "error": "MySQL module not available"}
    try:
        summary = db_mysql.get_link_summary_by_sheet()
        return {"ok": True, "summary": summary}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.post("/add_link")
def add_link_to_pool(
    url: str = Query(..., description="URL to add"),
    sheet: str = Query(..., description="Sheet name"),
    row: int = Query(..., description="Row number"),
    col: int = Query(1, description="Column number"),
):
    """
    Thêm link mới vào LINK_POOL và LINK_POOL_LIST.
    Format: { url, sheet, row, col, address, gid, sheet_name }
    """
    try:
        # Validate URL
        if not url or not url.startswith("http"):
            return {"ok": False, "error": "Invalid URL"}
        
        # Parse Google Sheets info if applicable
        gid = None
        sheet_name = None
        if "docs.google.com/spreadsheets" in url:
            file_id = extract_gsheets_file_id(url)
            parsed = urlparse(url)
            qs = parse_qs(parsed.fragment) if parsed.fragment else parse_qs(parsed.query)
            if "gid" in qs:
                gid = qs["gid"][0]
        
        # Create address (A1 notation)
        address = f"{chr(64 + col)}{row}"  # Simple: col 1 = A, col 2 = B, etc.
        
        # Create link record
        link_record = {
            "url": url,
            "sheet": sheet,
            "row": row,
            "col": col,
            "address": address,
            "gid": gid,
            "sheet_name": sheet_name,
        }
        
        # Add to LINK_POOL_LIST (flat list)
        LINK_POOL_LIST.append(link_record)
        
        # Add to LINK_POOL (map: url -> list of locations)
        if url not in LINK_POOL:
            LINK_POOL[url] = []
        LINK_POOL[url].append(link_record)
        
        # Add to LINK_POOL_MAP
        if url not in LINK_POOL_MAP:
            LINK_POOL_MAP[url] = []
        LINK_POOL_MAP[url].append(link_record)
        
        # Sync to MySQL if available
        if HAS_MYSQL:
            try:
                db_mysql.insert_links_batch([link_record])
            except Exception as e:
                _dlog(f"[add_link] Failed to sync to MySQL: {e}")
        
        _dlog(f"[add_link] Added: {url} at {sheet}!{address}")
        
        return {
            "ok": True,
            "link": link_record,
            "total_links": len(LINK_POOL_LIST),
            "message": f"Link added successfully to {sheet}!{address}"
        }
    
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.post("/mysql/sync_ctv_data")
def mysql_sync_ctv_data(clear_first: bool = False):
    """
    Đồng bộ DATABASE_ROWS (dữ liệu hoạt động) lên MySQL.
    Structure: [STT, MẢNG HOẠT ĐỘNG, ĐƠN VỊ, TÊN CHƯƠNG TRÌNH]
    """
    if not HAS_MYSQL:
        return {"ok": False, "error": "MySQL module not available"}
    
    try:
        if clear_first:
            deleted = db_mysql.clear_ctv_data_table()
            _dlog(f"[mysql] Cleared {deleted} old activity records")
        
        # Chuẩn bị dữ liệu từ DATABASE_ROWS
        ctv_records = []
        for row in DATABASE_ROWS:
            sheet = row.get("sheet", "")
            row_num = row.get("row", 0)
            cols = row.get("cols", [])
            text = row.get("text", "")
            links = row.get("links", [])
            
            # Structure thực tế: [STT, MẢNG HOẠT ĐỘNG, ĐƠN VỊ, TÊN CHƯƠNG TRÌNH]
            stt = ""
            category = ""  # Mảng hoạt động
            unit = ""      # Đơn vị
            program = ""   # Tên chương trình
            
            if len(cols) >= 1:
                stt = str(cols[0]).strip()
            if len(cols) >= 2:
                category = str(cols[1]).strip()
            if len(cols) >= 3:
                unit = str(cols[2]).strip()
            if len(cols) >= 4:
                program = str(cols[3]).strip()
            
            # Skip header rows và empty rows
            if stt.upper() in ["STT", "***", "DANH SÁCH", "THÀNH ĐOÀN", "BAN CHẤP HÀNH", "ĐẠI HỌC"]:
                continue
            if not category and not unit and not program:
                continue
            
            # Chuẩn hóa text cho search
            text_fixed = repair_text(text)
            text_normalized = fold_vi(text_fixed)
            
            # Sử dụng unit làm "full_name" cho search (tên đơn vị)
            # Và category làm "mssv" (mảng hoạt động)
            full_name = unit if unit else category
            name_normalized = fold_vi(full_name) if full_name else ""
            
            ctv_records.append({
                "sheet": sheet,
                "row": row_num,
                "full_name": full_name,  # Tên đơn vị
                "full_name_normalized": name_normalized,
                "mssv": stt,  # Số thứ tự (thay vì MSSV)
                "unit": category,  # Mảng hoạt động
                "program": program,  # Tên chương trình
                "row_text": text_fixed,
                "row_text_normalized": text_normalized,
                "links": links if isinstance(links, list) else [],
            })
        
        # Insert batch
        inserted = db_mysql.insert_ctv_data_batch(ctv_records)
        new_count = db_mysql.get_ctv_data_count()
        
        return {
            "ok": True,
            "inserted": inserted,
            "total_in_db": new_count,
            "synced_from_memory": len(DATABASE_ROWS),
            "structure": "STT | MẢNG HOẠT ĐỘNG | ĐƠN VỊ | TÊN CHƯƠNG TRÌNH",
            "hint": "Data structure: activities by units, not individual students"
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/mysql/ctv/count")
def mysql_ctv_count():
    """Đếm số CTV records trong MySQL."""
    if not HAS_MYSQL:
        return {"ok": False, "error": "MySQL module not available"}
    try:
        count = db_mysql.get_ctv_data_count()
        return {"ok": True, "count": count}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/mysql/ctv/search_name")
def mysql_ctv_search_name(q: str = Query(..., description="Tên để tìm"), limit: int = 50):
    """Tìm CTV theo tên (có dấu hoặc không dấu)."""
    if not HAS_MYSQL:
        return {"ok": False, "error": "MySQL module not available"}
    
    try:
        start_time = time.time()
        results = db_mysql.search_ctv_by_name(q, limit=limit)
        exec_time_ms = int((time.time() - start_time) * 1000)
        
        return {
            "ok": True,
            "query": q,
            "results": results,
            "count": len(results),
            "execution_time_ms": exec_time_ms,
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/mysql/ctv/search_mssv")
def mysql_ctv_search_mssv(mssv: str = Query(..., description="MSSV để tìm"), limit: int = 50):
    """Tìm CTV theo MSSV."""
    if not HAS_MYSQL:
        return {"ok": False, "error": "MySQL module not available"}
    
    try:
        start_time = time.time()
        results = db_mysql.search_ctv_by_mssv(mssv, limit=limit)
        exec_time_ms = int((time.time() - start_time) * 1000)
        
        return {
            "ok": True,
            "mssv": mssv,
            "results": results,
            "count": len(results),
            "execution_time_ms": exec_time_ms,
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/mysql/ctv/search")
def mysql_ctv_search(q: str = Query(..., description="Tìm kiếm tổng hợp"), limit: int = 50):
    """
    Tìm kiếm CTV theo tên, MSSV, đơn vị, hoặc chương trình.
    Tự động detect: nếu là số → search MSSV, nếu là text → search full-text.
    """
    if not HAS_MYSQL:
        return {"ok": False, "error": "MySQL module not available"}
    
    try:
        start_time = time.time()
        
        # Detect query type
        if q.strip().isdigit():
            # Search by MSSV
            results = db_mysql.search_ctv_by_mssv(q, limit=limit)
            search_type = "mssv"
        else:
            # Full-text search
            results = db_mysql.search_ctv_fulltext(q, limit=limit)
            search_type = "fulltext"
        
        exec_time_ms = int((time.time() - start_time) * 1000)
        
        return {
            "ok": True,
            "query": q,
            "search_type": search_type,
            "results": results,
            "count": len(results),
            "execution_time_ms": exec_time_ms,
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/mysql/ctv/by_sheet")
def mysql_ctv_by_sheet(sheet: str = Query(..., description="Sheet name"), limit: int = 100):
    """Lấy tất cả CTV trong một sheet."""
    if not HAS_MYSQL:
        return {"ok": False, "error": "MySQL module not available"}
    
    try:
        results = db_mysql.get_ctv_by_sheet(sheet, limit=limit)
        return {
            "ok": True,
            "sheet": sheet,
            "results": results,
            "count": len(results),
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/debug/sample_rows")
def debug_sample_rows(limit: int = 5):
    """
    Debug endpoint: Xem structure của DATABASE_ROWS để điều chỉnh parsing logic.
    """
    sample = []
    for i, row in enumerate(DATABASE_ROWS[:limit]):
        sample.append({
            "index": i,
            "sheet": row.get("sheet", ""),
            "row": row.get("row", 0),
            "cols": row.get("cols", []),
            "text": row.get("text", "")[:200],  # Truncate text
            "links": row.get("links", []),
            "num_cols": len(row.get("cols", [])),
        })
    return {
        "ok": True,
        "total_rows": len(DATABASE_ROWS),
        "sample": sample,
        "hint": "Check 'cols' array to identify which column contains name, MSSV, unit, program"
    }

@app.post("/mysql/sync_content")
async def mysql_sync_content(
    url: str = Query(..., description="URL to fetch and sync"),
    gid: Optional[str] = Query(None, description="Google Sheet GID if applicable")
):
    """
    Fetch nội dung từ URL và lưu vào MySQL content database.
    """
    if not HAS_MYSQL:
        return {"ok": False, "error": "MySQL module not available"}
    
    try:
        # Fetch content
        text = await fetch_text_from_url(url)
        
        if not text:
            # Fetch failed
            db_mysql.upsert_fetched_content(
                url=url,
                raw_content=None,
                normalized_content=None,
                content_type="unknown",
                gid=gid,
                row_count=0,
                status="error",
                error_message="Failed to fetch content"
            )
            return {"ok": False, "error": "Failed to fetch content", "url": url}
        
        # Parse content
        normalized = repair_text(text)
        
        # Nếu là CSV/Sheets, parse thành rows
        rows_data = []
        row_count = 0
        try:
            if is_google_sheets_url(url) or url.endswith(".csv") or "format=csv" in url:
                csv_rows = read_csv_text(text)
                row_count = len(csv_rows)
                
                # Prepare parsed rows data
                for idx, row in enumerate(csv_rows[:1000], start=1):  # Limit 1000 rows
                    row_text = " | ".join("" if c is None else str(c) for c in row)
                    rows_data.append({
                        "row_number": idx,
                        "values": row,
                        "text": row_text,
                        "normalized": fold_vi(row_text),
                    })
        except Exception as e:
            _dlog(f"[mysql] Failed to parse CSV: {e}")
        
        # Upsert main content
        db_mysql.upsert_fetched_content(
            url=url,
            raw_content=text[:1000000],  # Limit to 1MB
            normalized_content=normalized[:1000000],
            content_type="text/csv" if row_count > 0 else "text/plain",
            gid=gid,
            row_count=row_count,
            status="ok",
            error_message=None
        )
        
        # Insert parsed rows
        inserted_rows = 0
        if rows_data:
            inserted_rows = db_mysql.insert_parsed_rows_batch(url, rows_data)
        
        return {
            "ok": True,
            "url": url,
            "row_count": row_count,
            "parsed_rows_inserted": inserted_rows,
            "content_length": len(text),
        }
    
    except Exception as e:
        return {"ok": False, "error": str(e), "url": url}

@app.post("/mysql/sync_all_content")
async def mysql_sync_all_content(limit: int = 50):
    """
    Fetch và sync nội dung của tất cả links trong LINK_POOL (hoặc limit đầu tiên).
    """
    if not HAS_MYSQL:
        return {"ok": False, "error": "MySQL module not available"}
    
    try:
        results = {"ok": True, "processed": 0, "success": 0, "failed": 0, "details": []}
        
        links_to_process = _iter_all_indexed_links(limit=limit)
        
        for item in links_to_process:
            url = item.get("url")
            gid = item.get("gid")
            
            if not url:
                continue
            
            try:
                # Fetch
                text = await fetch_text_from_url(url)
                
                if text:
                    # Parse and insert
                    normalized = repair_text(text)
                    row_count = 0
                    
                    try:
                        if is_google_sheets_url(url) or url.endswith(".csv"):
                            csv_rows = read_csv_text(text)
                            row_count = len(csv_rows)
                            
                            rows_data = []
                            for idx, row in enumerate(csv_rows[:500], start=1):
                                row_text = " | ".join("" if c is None else str(c) for c in row)
                                rows_data.append({
                                    "row_number": idx,
                                    "values": row,
                                    "text": row_text,
                                    "normalized": fold_vi(row_text),
                                })
                            
                            if rows_data:
                                db_mysql.insert_parsed_rows_batch(url, rows_data)
                    except Exception:
                        pass
                    
                    db_mysql.upsert_fetched_content(
                        url=url,
                        raw_content=text[:500000],
                        normalized_content=normalized[:500000],
                        content_type="text/csv" if row_count > 0 else "text/plain",
                        gid=gid,
                        row_count=row_count,
                        status="ok",
                        error_message=None
                    )
                    
                    results["success"] += 1
                    results["details"].append({"url": url, "status": "ok", "rows": row_count})
                else:
                    db_mysql.upsert_fetched_content(
                        url=url, raw_content=None, normalized_content=None,
                        gid=gid, status="error", error_message="fetch_failed"
                    )
                    results["failed"] += 1
                    results["details"].append({"url": url, "status": "failed"})
                
                results["processed"] += 1
                
            except Exception as e:
                results["failed"] += 1
                results["details"].append({"url": url, "status": "error", "error": str(e)})
        
        return results
    
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/mysql/content/summary")
def mysql_content_summary():
    """Thống kê content đã fetch."""
    if not HAS_MYSQL:
        return {"ok": False, "error": "MySQL module not available"}
    try:
        summary = db_mysql.get_content_summary()
        return {"ok": True, "summary": summary}
    except Exception as e:
        return {"ok": False, "error": str(e)}

@app.get("/mysql/search")
def mysql_search(q: str = Query(..., description="Search query"), limit: int = 50):
    """Full-text search trong MySQL parsed_rows table."""
    if not HAS_MYSQL:
        return {"ok": False, "error": "MySQL module not available"}
    
    try:
        start_time = time.time()
        q_fixed, q_fold = normalize_query(q)
        
        results = db_mysql.search_in_parsed_rows(q_fixed, limit=limit)
        
        # Fallback: search với query không dấu
        if not results:
            results = db_mysql.search_in_parsed_rows(q_fold, limit=limit)
        
        exec_time_ms = int((time.time() - start_time) * 1000)
        
        # Log query
        db_mysql.log_search_query(q, q_fold, len(results), exec_time_ms)
        
        return {
            "ok": True,
            "query": q,
            "results": results,
            "count": len(results),
            "execution_time_ms": exec_time_ms,
        }
    except Exception as e:
        return {"ok": False, "error": str(e)}

# =========================
# Startup
# =========================
@app.on_event("startup")
def _startup_build():
    global DATABASE_ROWS, SHEETS
    _dlog(f"Building database... (deep={'ON' if USE_DEEP else 'OFF'})")
    rows, sheets = build_database(verbose=False, deep=USE_DEEP)
    DATABASE_ROWS = rows
    SHEETS = sheets
    _dlog(f"Database built with {len(DATABASE_ROWS)} rows from {len(SHEETS)} sheets; {len(LINK_POOL)} links")
